from __future__ import annotations

import csv
from pathlib import Path
from typing import Iterable

from openpyxl import Workbook, load_workbook


def read_table(
    path: str | Path, *, delimiter: str = "\t", sheet: str | int = 0
) -> tuple[list[str], list[dict[str, object]]]:
    table_path = Path(path)
    if table_path.suffix.lower() == ".xlsx":
        workbook = load_workbook(table_path, read_only=True, data_only=True)
        worksheet = (
            workbook.worksheets[sheet]
            if isinstance(sheet, int)
            else workbook[sheet]
        )
        iterator = worksheet.iter_rows(values_only=True)
        try:
            header = [str(value) if value is not None else "" for value in next(iterator)]
        except StopIteration as exc:
            raise ValueError("Workbook sheet is empty") from exc
        if any(not column for column in header) or len(set(header)) != len(header):
            raise ValueError("Table header contains empty or duplicate columns")
        rows = [dict(zip(header, values, strict=False)) for values in iterator]
        workbook.close()
        return header, rows

    with table_path.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle, delimiter=delimiter)
        if not reader.fieldnames:
            raise ValueError("Delimited table has no header")
        header = list(reader.fieldnames)
        if any(not column for column in header) or len(set(header)) != len(header):
            raise ValueError("Table header contains empty or duplicate columns")
        return header, list(reader)


def write_table(
    path: str | Path,
    header: Iterable[str],
    rows: Iterable[dict[str, object]],
    *,
    delimiter: str = "\t",
) -> Path:
    table_path = Path(path)
    table_path.parent.mkdir(parents=True, exist_ok=True)
    columns = list(header)
    row_list = list(rows)
    if table_path.suffix.lower() == ".xlsx":
        workbook = Workbook(write_only=True)
        worksheet = workbook.create_sheet("normalized")
        worksheet.append(columns)
        for row in row_list:
            worksheet.append([row.get(column) for column in columns])
        workbook.save(table_path)
        return table_path

    with table_path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(
            handle, fieldnames=columns, delimiter=delimiter, lineterminator="\n"
        )
        writer.writeheader()
        writer.writerows(row_list)
    return table_path


def normalize_hgvs_table(
    input_path: str | Path,
    output_path: str | Path,
    *,
    coding_columns: tuple[str, ...] = ("HGVSc",),
    protein_columns: tuple[str, ...] = ("HGVSp", "HGVSp_short"),
    delimiter: str = "\t",
    sheet: str | int = 0,
) -> dict[str, int]:
    from .hgvs import normalize_hgvs

    header, rows = read_table(input_path, delimiter=delimiter, sheet=sheet)
    requested = (*coding_columns, *protein_columns)
    missing = [column for column in requested if column not in header]
    if missing:
        raise ValueError(f"HGVS table is missing columns: {', '.join(missing)}")
    output_header = list(header)
    for column in requested:
        output_header.extend(
            [
                f"{column}_before",
                f"{column}_changed",
                f"{column}_change_reason",
                f"{column}_syntax_status",
            ]
        )
    changed_cells = 0
    invalid_syntax_cells = 0
    for row in rows:
        for column in coding_columns:
            value = row.get(column)
            body = str(value or "").strip().rsplit(":", 1)[-1].casefold()
            result = normalize_hgvs(
                value, kind="n" if body.startswith("n.") else "c"
            )
            row[f"{column}_before"] = row.get(column)
            row[column] = result.normalized
            row[f"{column}_changed"] = result.changed
            row[f"{column}_change_reason"] = ";".join(result.reasons)
            row[f"{column}_syntax_status"] = result.syntax_status
            changed_cells += int(result.changed)
            invalid_syntax_cells += int(result.syntax_status == "unvalidated-syntax")
        for column in protein_columns:
            result = normalize_hgvs(row.get(column), kind="p")
            row[f"{column}_before"] = row.get(column)
            row[column] = result.normalized
            row[f"{column}_changed"] = result.changed
            row[f"{column}_change_reason"] = ";".join(result.reasons)
            row[f"{column}_syntax_status"] = result.syntax_status
            changed_cells += int(result.changed)
            invalid_syntax_cells += int(result.syntax_status == "unvalidated-syntax")
    write_table(output_path, output_header, rows, delimiter=delimiter)
    return {
        "rows": len(rows),
        "changed_cells": changed_cells,
        "unvalidated_syntax_cells": invalid_syntax_cells,
    }
